import os
import re
import shutil
import openpyxl
from openpyxl.styles import PatternFill

# 加载规则文件
def load_rules(file_path):
    """加载规则段（强制小写验证）"""
    try:
        with open(file_path, 'r') as f:
            raw_rules = {line.strip().lower() for line in f if line.strip()}  # 规则强制小写
            
        mandatory_rules = {'m100-1.2w', 'f1w', 'f2w', 'f3w', 'f4w', 'f5w', 'f6w', 'm100-8w', 'm100-9w'}
        missing = mandatory_rules - raw_rules
        if missing:
            raise ValueError(f"规则文件缺少必须条目：{', '.join(missing)}")
        return raw_rules
    except Exception as e:
        print(f"规则加载失败：{str(e)}")
        exit(1)

# 移动SKU文件夹并清理空目录
def move_and_cleanup_sku_folders(base_folder):
    """移动 SKU 文件夹，并清理空文件夹"""
    if not os.path.exists(base_folder):
        print(f"目标文件夹 '{base_folder}' 不存在！")
        return

    for root, dirs, files in os.walk(base_folder, topdown=False):
        for file in files:
            if file.endswith(".jpg"):
                file_path = os.path.join(root, file)
                sku_folder = os.path.basename(os.path.dirname(file_path))
                dest_folder = os.path.join(base_folder, sku_folder)
                dest_path = os.path.join(dest_folder, file)

                if not os.path.exists(dest_folder):
                    os.makedirs(dest_folder)

                try:
                    shutil.move(file_path, dest_path)
                    print(f"文件移动成功：'{file_path}' -> '{dest_path}'")
                except Exception as e:
                    print(f"文件移动失败：'{file_path}'，错误：{e}")

        if not os.listdir(root):  # 如果当前目录为空，删除它
            try:
                os.rmdir(root)
                print(f"删除空文件夹：'{root}'")
            except Exception as e:
                print(f"删除失败：'{root}'，错误：{e}")

# 修改文件夹名称
def rename_subfolders(folder_path):
    """按规则修正SKU文件夹命名"""
    if not os.path.exists(folder_path):
        print(f"目标文件夹 '{folder_path}' 不存在！")
        return

    for subfolder_name in os.listdir(folder_path):
        subfolder_path = os.path.join(folder_path, subfolder_name)
        if not os.path.isdir(subfolder_path):
            continue

        name_length = len(subfolder_name)
        if name_length == 19 or name_length == 22:
            new_name = subfolder_name[:-2]
            new_path = os.path.join(folder_path, new_name)
            try:
                os.rename(subfolder_path, new_path)
                print(f"重命名：'{subfolder_name}' -> '{new_name}'")
            except Exception as e:
                print(f"重命名失败：'{subfolder_name}' -> '{new_name}'，错误：{e}")
        else:
            print(f"跳过：'{subfolder_name}'，长度不符合要求")

# 检查图片文件命名并生成报告
def check_and_fix_folders(root_folder, rule_file):
    rules = load_rules(rule_file)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "严格小写验证报告"
    sheet.append(["文件夹", "问题文件", "错误类型", "正确示例"])

    yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = yellow_fill

    errors_found = False

    file_pattern = re.compile(
        r'^[a-z\-]+-'    # 关键词段（必须小写）
        r'([a-z0-9.\-]+)'  # 规则段（必须小写）
        r'\.jpg$'        # 扩展名（必须小写）
    )

    for folder in os.listdir(root_folder):
        folder_path = os.path.join(root_folder, folder)
        if not os.path.isdir(folder_path):
            continue

        errors = []
        all_files = os.listdir(folder_path)
        found_rules = set()

        # 判断是否是有效的 SKU 文件夹
        # 如果文件夹命名长度是 20 位或 17 位，则认为是正确的命名
        if len(folder) not in (20, 17):
            errors.append(("-", "文件夹命名错误", "正确命名长度应为20位或17位"))

        # 文件检查
        for file in all_files:
            if not file.endswith('.jpg'):
                if file.lower().endswith('.jpg'):
                    errors.append((file, "扩展名大小写错误", "必须使用小写.jpg"))
                continue

            if file != file.lower():
                errors.append((file, "文件名大小写错误", "所有字母必须小写，如：pipe-shelves-m100-1.2w.jpg"))
                continue

            match = file_pattern.match(file)
            if not match:
                errors.append((file, "命名结构错误", "示例：pipe-shelves-m100-1.2w.jpg"))
                continue

            rule_segment = match.group(1)
            if rule_segment not in rules:
                errors.append((file, "未授权规则段", f"允许的规则如：m100-1.2w"))
            else:
                found_rules.add(rule_segment)

        mandatory_check = {'m100-1.2w', 'f1w', 'f2w', 'f3w', 'f4w', 'f5w', 'f6w', 'm100-8w', 'm100-9w'}
        missing = mandatory_check - found_rules
        if missing:
            errors.append((
                ", ".join(missing),
                "缺少核心文件",
                "必须存在的规则段（可带小写关键词段）"
            ))

        if errors:
            errors_found = True
            for error in errors:
                sheet.append([folder, error[0], error[1], error[2]])
            os.rename(folder_path, os.path.join(root_folder, f"INVALID_{folder}"))

    if errors_found:
        report_path = os.path.join(root_folder, "图片命名规则验证报告.xlsx")
        wb.save(report_path)
        print(f"发现异常，报告已生成：{report_path}")
    else:
        print("未发现异常")

# 主程序
if __name__ == "__main__":
    desktop_path = os.path.expanduser("~/Desktop")
    img_folder = os.path.join(desktop_path, "ImgNameCheck")
    rule_file_path = os.path.join(desktop_path, "withoutLogo_ImgName_rules.txt")

    move_and_cleanup_sku_folders(img_folder)
    rename_subfolders(img_folder)
    check_and_fix_folders(img_folder, rule_file_path)
